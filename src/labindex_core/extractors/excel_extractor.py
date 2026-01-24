"""
Excel file extractor.

Handles .xlsx, .xls files using openpyxl and xlrd.
"""

from pathlib import Path
from typing import List

from .base import TextExtractor, ExtractionResult


class ExcelExtractor(TextExtractor):
    """Extract text from Excel files."""

    EXTENSIONS = ['.xlsx', '.xls']

    # Limits to avoid slow extraction on large data files
    MAX_SHEETS = 5          # Only process first N sheets
    MAX_ROWS_PER_SHEET = 200  # Only read first N rows per sheet
    MAX_CELLS_TOTAL = 5000   # Stop after this many cells total

    def _extract_impl(self, path: Path) -> ExtractionResult:
        """Extract text from all cells in all sheets."""
        ext = path.suffix.lower()

        if ext == '.xlsx':
            return self._extract_xlsx(path)
        elif ext == '.xls':
            return self._extract_xls(path)
        else:
            return ExtractionResult.failure(f"Unknown Excel format: {ext}")

    def _extract_xlsx(self, path: Path) -> ExtractionResult:
        """Extract from .xlsx using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ExtractionResult.failure("openpyxl not installed")

        try:
            # Load workbook (read-only for speed, data-only to get values not formulas)
            wb = load_workbook(path, read_only=True, data_only=True)

            all_text: List[str] = []
            sources = {}
            total_cells = 0
            hit_limit = False

            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                # Limit number of sheets
                if sheet_idx >= self.MAX_SHEETS:
                    hit_limit = True
                    break

                sheet = wb[sheet_name]
                sheet_text: List[str] = []
                row_count = 0

                for row in sheet.iter_rows():
                    # Limit rows per sheet
                    row_count += 1
                    if row_count > self.MAX_ROWS_PER_SHEET:
                        hit_limit = True
                        break

                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            value = str(cell.value).strip()
                            if value:
                                row_values.append(value)
                                total_cells += 1

                                # Limit total cells
                                if total_cells >= self.MAX_CELLS_TOTAL:
                                    hit_limit = True
                                    break

                    if row_values:
                        sheet_text.append(' '.join(row_values))

                    if hit_limit:
                        break

                if sheet_text:
                    sheet_content = '\n'.join(sheet_text)
                    all_text.append(f"[Sheet: {sheet_name}]\n{sheet_content}")
                    sources[sheet_name] = sheet_content[:500]

                if hit_limit:
                    break

            wb.close()

            return ExtractionResult(
                text='\n\n'.join(all_text),
                metadata={'sheets': list(sources.keys()), 'truncated': hit_limit},
                sources=sources
            )

        except Exception as e:
            return ExtractionResult.failure(f"Excel read error: {e}")

    def _extract_xls(self, path: Path) -> ExtractionResult:
        """Extract from .xls using xlrd."""
        try:
            import xlrd
        except ImportError:
            return ExtractionResult.failure("xlrd not installed (needed for .xls files)")

        try:
            wb = xlrd.open_workbook(path)

            all_text: List[str] = []
            sources = {}
            total_cells = 0
            hit_limit = False

            for sheet_idx, sheet_name in enumerate(wb.sheet_names()):
                # Limit number of sheets
                if sheet_idx >= self.MAX_SHEETS:
                    hit_limit = True
                    break

                sheet = wb.sheet_by_name(sheet_name)
                sheet_text: List[str] = []

                max_rows = min(sheet.nrows, self.MAX_ROWS_PER_SHEET)
                for row_idx in range(max_rows):
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.value:
                            value = str(cell.value).strip()
                            if value:
                                row_values.append(value)
                                total_cells += 1

                                if total_cells >= self.MAX_CELLS_TOTAL:
                                    hit_limit = True
                                    break

                    if row_values:
                        sheet_text.append(' '.join(row_values))

                    if hit_limit:
                        break

                if sheet_text:
                    sheet_content = '\n'.join(sheet_text)
                    all_text.append(f"[Sheet: {sheet_name}]\n{sheet_content}")
                    sources[sheet_name] = sheet_content[:500]

                if hit_limit:
                    break

            return ExtractionResult(
                text='\n\n'.join(all_text),
                metadata={'sheets': list(sources.keys()), 'truncated': hit_limit},
                sources=sources
            )

        except Exception as e:
            return ExtractionResult.failure(f"XLS read error: {e}")
